"""PDF report generation — SSP and POA&M via ReportLab."""
from __future__ import annotations

import io
from datetime import datetime, date

import asyncpg
from minio import Minio
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

from app.config import settings as _settings
from app.services.minio_service import upload_bytes, get_proxy_download_url, ensure_bucket

REPORTS_BUCKET = "cmmc-reports"

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _styles():
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="SectionHeading",
            parent=styles["Heading1"],
            fontSize=14,
            spaceAfter=6,
            spaceBefore=12,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SubHeading",
            parent=styles["Heading2"],
            fontSize=11,
            spaceAfter=4,
            spaceBefore=8,
        )
    )
    styles.add(
        ParagraphStyle(
            name="BodySmall",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
        )
    )
    return styles


def _add_page_number(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(
        letter[0] - 0.5 * inch,
        0.5 * inch,
        f"Page {doc.page}",
    )
    canvas.restoreState()


def _cell(text: str | None, style) -> Paragraph:
    return Paragraph(str(text or ""), style)


# ---------------------------------------------------------------------------
# SSP PDF
# ---------------------------------------------------------------------------


async def generate_ssp_pdf(
    program_id: str,
    conn: asyncpg.Connection,
    minio_client: Minio,
    minio_public: Minio | None = None,
) -> str:
    """Generate an SSP PDF and upload it to MinIO. Returns presigned download URL."""
    # ---- fetch program + org ------------------------------------------------
    program = await conn.fetchrow(
        """
        SELECT p.*, o.name AS org_name, o.cage_code AS org_cage_code
        FROM programs p
        JOIN orgs o ON p.org_id = o.id
        WHERE p.id = $1
        """,
        program_id,
    )
    if not program:
        raise ValueError(f"Program {program_id} not found")

    # ---- fetch controls (parent controls only, not objectives) ---------------
    controls = await conn.fetch(
        """
        SELECT
            cd.nist_id,
            cd.cmmc_id,
            cd.family,
            cd.requirement_text,
            pc.status,
            pc.implementation_notes
        FROM program_controls pc
        JOIN control_definitions cd ON pc.control_definition_id = cd.id
        WHERE pc.program_id = $1
          AND cd.is_objective = FALSE
          AND pc.is_applicable = TRUE
        ORDER BY cd.far_above_phase, cd.nist_sort_order
        """,
        program_id,
    )

    styles = _styles()
    story = []
    today = datetime.utcnow().strftime("%B %d, %Y")

    # ---- Title page ----------------------------------------------------------
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph("System Security Plan", styles["Title"]))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(f"Organization: {program['org_name']}", styles["Heading2"]))
    story.append(Paragraph(f"System: {program.get('system_name') or program['name']}", styles["Heading2"]))
    story.append(Paragraph(f"Date: {today}", styles["Normal"]))
    story.append(PageBreak())

    # ---- Section 1: System Overview ------------------------------------------
    story.append(Paragraph("1. System Overview", styles["SectionHeading"]))
    story.append(Paragraph(f"<b>System Name:</b> {program.get('system_name') or program['name']}", styles["Normal"]))
    cage_codes = program.get("cage_codes") or []
    if isinstance(cage_codes, str):
        cage_codes = [cage_codes]
    story.append(Paragraph(f"<b>CAGE Code(s):</b> {', '.join(cage_codes) if cage_codes else 'N/A'}", styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))
    desc = program.get("ssp_system_description") or "Not yet provided."
    story.append(Paragraph(f"<b>System Description:</b><br/>{desc}", styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    # ---- Section 2: Environment of Operation ---------------------------------
    story.append(Paragraph("2. Environment of Operation", styles["SectionHeading"]))
    env = program.get("ssp_environment_of_operation") or "Not yet provided."
    story.append(Paragraph(env, styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    # ---- Section 3: CUI Types ------------------------------------------------
    story.append(Paragraph("3. CUI Information Types", styles["SectionHeading"]))
    info_types = program.get("ssp_information_types") or "Not yet provided."
    story.append(Paragraph(info_types, styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    # ---- Section 4: Implementation Status ------------------------------------
    story.append(Paragraph("4. Control Implementation Status", styles["SectionHeading"]))
    cell_style = styles["BodySmall"]
    header_style = ParagraphStyle(
        name="TableHeader", parent=cell_style, textColor=colors.white,
        fontName="Helvetica-Bold", fontSize=8,
    )
    table_data = [[
        Paragraph(h, header_style)
        for h in ["NIST ID", "CMMC ID", "Family", "Status", "Implementation Notes"]
    ]]
    for row in controls:
        notes = row.get("implementation_notes") or ""
        table_data.append([
            _cell(row["nist_id"], cell_style),
            _cell(row.get("cmmc_id") or "", cell_style),
            _cell(row.get("family") or "", cell_style),
            _cell((row.get("status") or "").replace("_", " ").title(), cell_style),
            _cell(notes, cell_style),
        ])

    col_widths = [0.8 * inch, 0.8 * inch, 1.2 * inch, 1.1 * inch, 3.6 * inch]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f4f7")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ])
    )
    story.append(tbl)
    story.append(Spacer(1, 0.2 * inch))

    # ---- Section 5: Topology -------------------------------------------------
    story.append(Paragraph("5. System Topology", styles["SectionHeading"]))
    topo = program.get("topology_narrative") or "Not yet provided."
    story.append(Paragraph(topo, styles["Normal"]))
    diagram_url = program.get("topology_diagram_url")
    if diagram_url:
        story.append(Spacer(1, 0.1 * inch))
        story.append(Paragraph(f"Network diagram: {diagram_url}", styles["Normal"]))

    # ---- Build PDF ----------------------------------------------------------
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, rightMargin=0.75 * inch, leftMargin=0.75 * inch,
                            topMargin=1 * inch, bottomMargin=0.75 * inch)
    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    pdf_bytes = buf.getvalue()

    # ---- Upload & presign ---------------------------------------------------
    ensure_bucket(minio_client, REPORTS_BUCKET)
    key = f"{program_id}/ssp_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    upload_bytes(minio_client, REPORTS_BUCKET, key, pdf_bytes, "application/pdf")
    return get_proxy_download_url(REPORTS_BUCKET, key, _settings.api_url, _settings.webhook_secret)


# ---------------------------------------------------------------------------
# POA&M PDF
# ---------------------------------------------------------------------------


async def generate_poam_pdf(
    program_id: str,
    conn: asyncpg.Connection,
    minio_client: Minio,
    minio_public: Minio | None = None,
) -> str:
    """Generate a POA&M PDF and upload it to MinIO. Returns presigned download URL."""
    program = await conn.fetchrow(
        """
        SELECT p.*, o.name AS org_name
        FROM programs p
        JOIN orgs o ON p.org_id = o.id
        WHERE p.id = $1
        """,
        program_id,
    )
    if not program:
        raise ValueError(f"Program {program_id} not found")

    # Only non-fully-implemented applicable controls
    controls = await conn.fetch(
        """
        SELECT
            cd.nist_id,
            cd.cmmc_id,
            cd.requirement_text,
            pc.status,
            pc.implementation_notes,
            pc.target_completion_date
        FROM program_controls pc
        JOIN control_definitions cd ON pc.control_definition_id = cd.id
        WHERE pc.program_id = $1
          AND cd.is_objective = FALSE
          AND pc.is_applicable = TRUE
          AND pc.status != 'fully_implemented'
        ORDER BY cd.far_above_phase, cd.nist_sort_order
        """,
        program_id,
    )

    styles = _styles()
    story = []
    today = datetime.utcnow().strftime("%B %d, %Y")

    # ---- Title page ----------------------------------------------------------
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph("Plan of Action & Milestones", styles["Title"]))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(f"Organization: {program['org_name']}", styles["Heading2"]))
    story.append(Paragraph(f"System: {program.get('system_name') or program['name']}", styles["Heading2"]))
    story.append(Paragraph(f"Date: {today}", styles["Normal"]))
    story.append(PageBreak())

    # ---- POA&M Table ---------------------------------------------------------
    story.append(Paragraph("Open Items", styles["SectionHeading"]))
    if not controls:
        story.append(Paragraph("No open items — all applicable controls are fully implemented.", styles["Normal"]))
    else:
        cell_style = styles["BodySmall"]
        header_style = ParagraphStyle(
            name="TableHeader", parent=cell_style, textColor=colors.white,
            fontName="Helvetica-Bold", fontSize=8,
        )
        table_data = [[
            Paragraph(h, header_style)
            for h in ["Control ID", "Description", "Status", "Responsible Org", "Resources", "Milestone Date", "Remediation Plan"]
        ]]
        for row in controls:
            desc = row.get("requirement_text") or ""
            milestone = row.get("target_completion_date")
            milestone_str = milestone.strftime("%Y-%m-%d") if isinstance(milestone, date) else (str(milestone)[:10] if milestone else "TBD")
            notes = row.get("implementation_notes") or ""
            table_data.append([
                _cell(row["nist_id"], cell_style),
                _cell(desc, cell_style),
                _cell((row.get("status") or "").replace("_", " ").title(), cell_style),
                _cell(program["org_name"], cell_style),
                _cell("TBD", cell_style),
                _cell(milestone_str, cell_style),
                _cell(notes or "Remediation plan pending.", cell_style),
            ])

        # Landscape letter = 11" wide; 9.5" usable after margins
        col_widths = [0.65 * inch, 2.0 * inch, 1.0 * inch, 1.2 * inch, 0.75 * inch, 0.9 * inch, 2.95 * inch]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f4f7")]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ])
        )
        story.append(tbl)

    # ---- Build PDF ----------------------------------------------------------
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), rightMargin=0.75 * inch, leftMargin=0.75 * inch,
                            topMargin=1 * inch, bottomMargin=0.75 * inch)
    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    pdf_bytes = buf.getvalue()

    # ---- Upload & presign ---------------------------------------------------
    ensure_bucket(minio_client, REPORTS_BUCKET)
    key = f"{program_id}/poam_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    upload_bytes(minio_client, REPORTS_BUCKET, key, pdf_bytes, "application/pdf")
    return get_proxy_download_url(REPORTS_BUCKET, key, _settings.api_url, _settings.webhook_secret)


# ---------------------------------------------------------------------------
# SPRS Score Summary Excel
# ---------------------------------------------------------------------------

async def generate_sprs_xlsx(
    program_id: str,
    conn: asyncpg.Connection,
    minio_client: Minio,
    minio_public: Minio | None = None,
) -> str:
    """Generate a SPRS score breakdown Excel and upload to MinIO."""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter

    program = await conn.fetchrow(
        """
        SELECT p.*, o.name AS org_name, o.cage_code AS org_cage_code
        FROM programs p JOIN orgs o ON p.org_id = o.id
        WHERE p.id = $1
        """,
        program_id,
    )
    if not program:
        raise ValueError(f"Program {program_id} not found")

    controls = await conn.fetch(
        """
        SELECT
            cd.nist_id, cd.cmmc_id, cd.family, cd.family_abbrev,
            cd.requirement_text, cd.dod_score_value,
            pc.status, pc.is_applicable
        FROM program_controls pc
        JOIN control_definitions cd ON pc.control_definition_id = cd.id
        WHERE pc.program_id = $1 AND cd.is_objective = FALSE
        ORDER BY cd.family_abbrev, cd.nist_sort_order
        """,
        program_id,
    )

    # ── Score calculation (mirrors sprs_service.py logic) ──────────────────
    SPRS_MAX = 110
    sprs = SPRS_MAX
    ssp_ok = True
    for c in controls:
        if not c["is_applicable"]:
            continue
        if c["nist_id"] == "3.12.4" and c["status"] != "fully_implemented":
            ssp_ok = False
        if c["status"] != "fully_implemented" and c["dod_score_value"]:
            sprs -= c["dod_score_value"]
    if not ssp_ok:
        sprs = -203

    # ── Domain subtotals ────────────────────────────────────────────────────
    from collections import defaultdict
    domains: dict[str, dict] = defaultdict(lambda: {
        "name": "", "total": 0, "implemented": 0, "points_possible": 0, "points_lost": 0
    })
    DOMAIN_FULL = {
        "AC": "Access Control", "AT": "Awareness & Training",
        "AU": "Audit & Accountability", "CM": "Configuration Management",
        "IA": "Identification & Authentication", "IR": "Incident Response",
        "MA": "Maintenance", "MP": "Media Protection",
        "PE": "Physical Protection", "PS": "Personnel Security",
        "RA": "Risk Assessment", "CA": "Security Assessment",
        "SC": "System & Comm Protection", "SI": "System & Info Integrity",
    }
    for c in controls:
        if not c["is_applicable"]:
            continue
        abbrev = c["family_abbrev"] or "??"
        d = domains[abbrev]
        d["name"] = DOMAIN_FULL.get(abbrev, c["family"] or abbrev)
        d["total"] += 1
        d["points_possible"] += c["dod_score_value"] or 0
        if c["status"] == "fully_implemented":
            d["implemented"] += 1
        else:
            d["points_lost"] += c["dod_score_value"] or 0

    # ── Build workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Colour palette
    NAVY = "1E3A5F"
    BLUE = "2563EB"
    GREEN = "16A34A"
    AMBER = "D97706"
    RED_C = "DC2626"
    LTGRAY = "F1F5F9"
    WHITE = "FFFFFF"
    DARK_RED = "991B1B"

    def hdr_font(bold=True, size=11, color=WHITE):
        return Font(bold=bold, size=size, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="CBD5E1")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Sheet 1: SPRS Summary ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "SPRS Summary"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 36

    def ws1_row(label, value, label_bold=False, value_fmt=None, label_fill=None, value_fill=None, row_height=None):
        r = ws1.max_row + 1
        lc = ws1.cell(r, 1, label)
        vc = ws1.cell(r, 2, value)
        lc.alignment = left()
        vc.alignment = left()
        lc.border = border()
        vc.border = border()
        if label_bold:
            lc.font = Font(bold=True, size=11)
        if label_fill:
            lc.fill = fill(label_fill)
        if value_fill:
            vc.fill = fill(value_fill)
        if value_fmt:
            vc.number_format = value_fmt
        if row_height:
            ws1.row_dimensions[r].height = row_height
        return lc, vc

    # Title banner
    ws1.merge_cells("A1:B1")
    title_cell = ws1["A1"]
    cmmc_level = program.get("cmmc_level", 2)
    title_cell.value = f"DoD CMMC Level {cmmc_level} — {'SPRS Self-Assessment' if cmmc_level == 2 else 'Readiness'} Summary"
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = fill(NAVY)
    title_cell.alignment = center()
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:B2")
    sub_cell = ws1["A2"]
    sub_cell.value = "NIST SP 800-171 Rev 2 | Submit via PIEE Supplier Performance Risk System (SPRS)"
    sub_cell.font = Font(italic=True, size=9, color="475569")
    sub_cell.alignment = center()
    ws1.row_dimensions[2].height = 18

    ws1.append([])  # blank row 3

    today_str = datetime.utcnow().strftime("%B %d, %Y")
    ws1_row("Organization", program["org_name"] or "", label_bold=True)
    ws1_row("System Name", program.get("system_name") or program["name"] or "")
    ws1_row("CAGE Code", program.get("org_cage_code") or "")
    ws1_row("Assessment Date", today_str)
    ws1_row("Assessor", "Self-Assessment (Internal)")
    ws1.append([])

    # SPRS Score highlight
    score_color = GREEN if sprs >= 70 else (AMBER if sprs >= 0 else DARK_RED)
    r_sprs = ws1.max_row + 1
    lc = ws1.cell(r_sprs, 1, "SPRS SCORE")
    vc = ws1.cell(r_sprs, 2, sprs)
    lc.font = Font(bold=True, size=16, color=WHITE)
    vc.font = Font(bold=True, size=24, color=WHITE)
    lc.fill = fill(NAVY)
    vc.fill = fill(score_color)
    lc.alignment = center()
    vc.alignment = center()
    ws1.row_dimensions[r_sprs].height = 40

    ws1.append([])

    # Score interpretation
    interp = ("Acceptable — strong compliance posture" if sprs >= 70
              else ("Caution — partial compliance, remediation required" if sprs >= 0
                    else "Critical — SSP not implemented or significant gaps (floor: -203)"))
    ws1_row("Score Interpretation", interp, label_bold=True,
            value_fill=(GREEN if sprs >= 70 else (AMBER if sprs >= 0 else RED_C)))

    ws1_row("Maximum Possible Score", 110)
    ws1_row("Points Deducted", SPRS_MAX - sprs if not ssp_ok else SPRS_MAX - sprs)
    ws1_row("SSP Control (3.12.4) Implemented", "Yes" if ssp_ok else "No — SPRS floored to -203",
            value_fill=("F0FDF4" if ssp_ok else "FEF2F2"))

    ws1.append([])

    # PIEE submission instructions
    piee_label = ws1.cell(ws1.max_row + 1, 1, "PIEE Submission Instructions")
    piee_label.font = Font(bold=True, size=11, color=WHITE)
    piee_label.fill = fill(BLUE)
    piee_label.alignment = center()
    ws1.merge_cells(f"A{ws1.max_row}:B{ws1.max_row}")

    instructions = [
        ("Step 1", "Log in to PIEE at https://piee.eb.mil"),
        ("Step 2", "Navigate to: Supplier Performance Risk System (SPRS)"),
        ("Step 3", "Select 'NIST SP 800-171 DoD Assessment' → 'Add Assessment'"),
        ("Step 4", f"Enter your SPRS Score: {sprs}"),
        ("Step 5", "Enter the assessment date, system name, and CAGE code above"),
        ("Step 6", "Attach your SSP as supporting documentation"),
        ("Step 7", "Submit — score is visible to DoD contracting officers within 24h"),
        ("Note", "This self-assessment score is valid for 3 years per DFARS 252.204-7021"),
    ]
    for step, detail in instructions:
        lc, vc = ws1_row(step, detail)
        lc.fill = fill(LTGRAY)
        lc.font = Font(bold=True, size=10)
        vc.font = Font(size=10)

    # ── Sheet 2: Domain Breakdown ───────────────────────────────────────────
    ws2 = wb.create_sheet("Domain Breakdown")
    for col, (header, width) in enumerate([
        ("Domain", 8), ("Full Name", 30), ("Controls", 10),
        ("Implemented", 13), ("Pct", 9), ("Points Possible", 15), ("Points Lost", 13)
    ], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
        cell = ws2.cell(1, col, header)
        cell.font = hdr_font()
        cell.fill = fill(NAVY)
        cell.alignment = center()
        cell.border = border()
    ws2.row_dimensions[1].height = 22

    for r_idx, (abbrev, d) in enumerate(sorted(domains.items()), 2):
        pct = round(d["implemented"] / d["total"] * 100) if d["total"] else 0
        row_fill = fill("F0FDF4") if pct == 100 else (fill("FEF9C3") if pct >= 50 else fill("FEF2F2"))
        vals = [abbrev, d["name"], d["total"], d["implemented"],
                f"{pct}%", d["points_possible"], d["points_lost"]]
        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(r_idx, col_idx, val)
            cell.fill = row_fill
            cell.alignment = center() if col_idx != 2 else left()
            cell.border = border()
            cell.font = Font(size=10)

    # ── Sheet 3: Control Detail ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Control Detail")
    headers3 = ["NIST ID", "CMMC ID", "Domain", "Requirement (truncated)", "Status", "DoD Pts", "Deducted"]
    widths3 = [10, 14, 8, 52, 22, 9, 10]
    for col, (h, w) in enumerate(zip(headers3, widths3), 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
        cell = ws3.cell(1, col, h)
        cell.font = hdr_font()
        cell.fill = fill(NAVY)
        cell.alignment = center()
        cell.border = border()
    ws3.row_dimensions[1].height = 22
    ws3.freeze_panes = "A2"

    STATUS_FILLS = {
        "fully_implemented": "F0FDF4",
        "implementation_begun": "EFF6FF",
        "implementation_planned": "FEFCE8",
        "not_yet_addressed": "FEF2F2",
        "not_yet_assessed": "F8FAFC",
        "not_applicable": "F1F5F9",
    }
    for r_idx, c in enumerate(controls, 2):
        if not c["is_applicable"]:
            continue
        deducted = c["dod_score_value"] if c["status"] != "fully_implemented" else 0
        status_label = (c["status"] or "").replace("_", " ").title()
        req_text = c["requirement_text"] or ""
        req = req_text[:120] + ("…" if len(req_text) > 120 else "")
        row_fill = fill(STATUS_FILLS.get(c["status"] or "", "FFFFFF"))
        for col_idx, val in enumerate([
            c["nist_id"], c["cmmc_id"] or "", c["family_abbrev"] or "",
            req, status_label, c["dod_score_value"] or 0, deducted
        ], 1):
            cell = ws3.cell(r_idx, col_idx, val)
            cell.fill = row_fill
            cell.alignment = center() if col_idx not in (4, 5) else left()
            cell.border = border()
            cell.font = Font(size=9)
            if col_idx == 7 and deducted:
                cell.font = Font(size=9, color=DARK_RED, bold=True)

    # ── Save & upload ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    ensure_bucket(minio_client, REPORTS_BUCKET)
    key = f"{program_id}/sprs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    upload_bytes(minio_client, REPORTS_BUCKET, key, xlsx_bytes,
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return get_proxy_download_url(REPORTS_BUCKET, key, _settings.api_url, _settings.webhook_secret)


# ---------------------------------------------------------------------------
# Bundled Audit Package ZIP
# ---------------------------------------------------------------------------

EXPORTS_BUCKET = "cmmc-exports"


async def generate_audit_package_zip(
    program_id: str,
    conn: asyncpg.Connection,
    minio_client: Minio,
    minio_public: Minio | None = None,
) -> str:
    """Bundle SSP, POA&M, SPRS xlsx, artifact evidence, and manifest into a ZIP."""
    import zipfile
    from app.services.minio_service import download_bytes

    program = await conn.fetchrow(
        """
        SELECT p.*, o.name AS org_name, o.cage_code AS org_cage_code
        FROM programs p JOIN orgs o ON p.org_id = o.id
        WHERE p.id = $1
        """,
        program_id,
    )
    if not program:
        raise ValueError(f"Program {program_id} not found")

    org_name = (program["org_name"] or "org").replace(" ", "_")
    today = datetime.utcnow().strftime("%Y%m%d")
    zip_name = f"CMMC_AuditPackage_{org_name}_{today}.zip"

    # Fetch all artifact records with control info
    artifacts = await conn.fetch(
        """
        SELECT
            a.id, a.file_name, a.minio_key, a.mime_type,
            cd.nist_id, cd.family_abbrev,
            pc.status,
            ass.verdict, ass.confidence, ass.rationale
        FROM artifacts a
        JOIN program_controls pc ON a.program_control_id = pc.id
        JOIN control_definitions cd ON pc.control_definition_id = cd.id
        LEFT JOIN LATERAL (
            SELECT verdict, confidence, rationale
            FROM assessments WHERE artifact_id = a.id
            ORDER BY created_at DESC LIMIT 1
        ) ass ON TRUE
        WHERE pc.program_id = $1
        ORDER BY cd.family_abbrev, cd.nist_id, a.created_at
        """,
        program_id,
    )

    # Fetch all controls for manifest
    controls = await conn.fetch(
        """
        SELECT
            cd.nist_id, cd.cmmc_id, cd.family, cd.family_abbrev,
            cd.dod_score_value, pc.status, pc.is_applicable,
            pc.implementation_notes, pc.target_completion_date
        FROM program_controls pc
        JOIN control_definitions cd ON pc.control_definition_id = cd.id
        WHERE pc.program_id = $1 AND cd.is_objective = FALSE
        ORDER BY cd.family_abbrev, cd.nist_sort_order
        """,
        program_id,
    )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:

        # ── README ──────────────────────────────────────────────────────────
        _lvl = program.get("cmmc_level", 2)
        _ctrl_count = 110 if _lvl == 2 else 145
        readme = f"""CMMC Level {_lvl} Audit Package
Organization: {program['org_name']}
System: {program.get('system_name') or program['name']}
CAGE Code: {program.get('org_cage_code') or 'N/A'}
Generated: {datetime.utcnow().strftime('%B %d, %Y at %H:%M UTC')}

CONTENTS
--------
SSP.pdf                  — System Security Plan (all {_ctrl_count} controls)
POAM.pdf                 — Plan of Action & Milestones (open items)
SPRS_Summary.xlsx        — SPRS score breakdown (submit score to PIEE)
manifest.csv             — Full control inventory with status and evidence
evidence/                — Uploaded evidence files, organized by domain/control

HOW TO USE
----------
1. Provide SSP.pdf and POAM.pdf to your C3PAO assessor as primary inputs.
2. Evidence files in evidence/ support each control assessment objective.
3. Use SPRS_Summary.xlsx to submit your self-assessed SPRS score to PIEE at https://piee.eb.mil
4. manifest.csv gives the assessor a quick view of your control implementation posture.

IMPORTANT: Self-assessed SPRS scores are valid for 3 years per DFARS 252.204-7021.
C3PAO certification supersedes self-assessment scores.
"""
        zf.writestr("README.txt", readme)

        # ── Manifest CSV ────────────────────────────────────────────────────
        import csv
        import io as _io
        csv_buf = _io.StringIO()
        writer = csv.writer(csv_buf)
        writer.writerow([
            "NIST ID", "CMMC ID", "Domain", "Status", "Applicable",
            "DoD Points", "Deducted", "Target Date", "Notes (truncated)"
        ])
        for c in controls:
            deducted = (c["dod_score_value"] or 0) if c["status"] != "fully_implemented" and c["is_applicable"] else 0
            notes = (c["implementation_notes"] or "")[:100]
            target = str(c["target_completion_date"])[:10] if c["target_completion_date"] else ""
            writer.writerow([
                c["nist_id"], c["cmmc_id"] or "", c["family_abbrev"] or "",
                c["status"] or "", "Yes" if c["is_applicable"] else "No",
                c["dod_score_value"] or 0, deducted, target, notes,
            ])
        zf.writestr("manifest.csv", csv_buf.getvalue())

        # ── Generate sub-documents (SSP, POA&M, SPRS) ───────────────────────
        ssp_bytes = await _generate_ssp_bytes(program_id, program, controls)
        zf.writestr("SSP.pdf", ssp_bytes)

        poam_bytes = await _generate_poam_bytes(program_id, program, controls)
        zf.writestr("POAM.pdf", poam_bytes)

        sprs_bytes = await _generate_sprs_bytes(program_id, program, controls)
        zf.writestr("SPRS_Summary.xlsx", sprs_bytes)

        # ── Evidence files ───────────────────────────────────────────────────
        for art in artifacts:
            if not art["minio_key"]:
                continue
            try:
                file_bytes = download_bytes(minio_client, "cmmc-artifacts", art["minio_key"])
            except Exception:
                continue  # skip unavailable artifacts gracefully
            domain = art["family_abbrev"] or "OTHER"
            nist = (art["nist_id"] or "unknown").replace(".", "_")
            fname = art["file_name"] or f"artifact_{art['id']}"
            zip_path = f"evidence/{domain}/{nist}/{fname}"
            zf.writestr(zip_path, file_bytes)

    zip_bytes = buf.getvalue()

    ensure_bucket(minio_client, EXPORTS_BUCKET)
    key = f"{program_id}/audit_package_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"
    upload_bytes(minio_client, EXPORTS_BUCKET, key, zip_bytes, "application/zip")
    return get_proxy_download_url(EXPORTS_BUCKET, key, _settings.api_url, _settings.webhook_secret)


# ── Internal byte-only helpers (no MinIO upload, just bytes) ────────────────

async def _generate_ssp_bytes(program_id: str, program, controls) -> bytes:
    """Build SSP PDF bytes without uploading to MinIO."""
    from reportlab.lib import colors as _colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    styles = _styles()
    story = []
    today = datetime.utcnow().strftime("%B %d, %Y")
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph("System Security Plan", styles["Title"]))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(f"Organization: {program['org_name']}", styles["Heading2"]))
    story.append(Paragraph(f"System: {program.get('system_name') or program['name']}", styles["Heading2"]))
    story.append(Paragraph(f"Date: {today}", styles["Normal"]))
    story.append(PageBreak())
    story.append(Paragraph("Control Implementation Status", styles["SectionHeading"]))

    cell_style = styles["BodySmall"]
    hdr_sty = ParagraphStyle(name="_H", parent=cell_style, textColor=_colors.white,
                              fontName="Helvetica-Bold", fontSize=8)
    tdata = [[Paragraph(h, hdr_sty) for h in ["NIST ID", "CMMC ID", "Family", "Status", "Notes"]]]
    for c in controls:
        if not c["is_applicable"]:
            continue
        tdata.append([
            Paragraph(c["nist_id"], cell_style),
            Paragraph(c.get("cmmc_id") or "", cell_style),
            Paragraph(c.get("family") or "", cell_style),
            Paragraph((c.get("status") or "").replace("_", " ").title(), cell_style),
            Paragraph((c.get("implementation_notes") or "")[:200], cell_style),
        ])
    col_w = [0.8 * inch, 0.8 * inch, 1.2 * inch, 1.1 * inch, 3.6 * inch]
    tbl = Table(tdata, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), _colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_colors.white, _colors.HexColor("#f2f4f7")]),
        ("GRID", (0, 0), (-1, -1), 0.3, _colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            rightMargin=0.75 * inch, leftMargin=0.75 * inch,
                            topMargin=1 * inch, bottomMargin=0.75 * inch)
    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    return buf.getvalue()


async def _generate_poam_bytes(program_id: str, program, controls) -> bytes:
    """Build POA&M PDF bytes without uploading to MinIO."""
    from reportlab.lib import colors as _colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    styles = _styles()
    story = []
    today = datetime.utcnow().strftime("%B %d, %Y")
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph("Plan of Action & Milestones", styles["Title"]))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(f"Organization: {program['org_name']}", styles["Heading2"]))
    story.append(Paragraph(f"Date: {today}", styles["Normal"]))
    story.append(PageBreak())

    open_controls = [c for c in controls if c["is_applicable"] and c["status"] != "fully_implemented"]
    if not open_controls:
        story.append(Paragraph("No open items — all applicable controls are fully implemented.", styles["Normal"]))
    else:
        cell_style = styles["BodySmall"]
        hdr_sty = ParagraphStyle(name="_HP", parent=cell_style, textColor=_colors.white,
                                  fontName="Helvetica-Bold", fontSize=8)
        tdata = [[Paragraph(h, hdr_sty) for h in
                  ["Control ID", "Description", "Status", "Responsible Org", "Resources", "Milestone", "Plan"]]]
        for c in open_controls:
            m = c.get("target_completion_date")
            ms = m.strftime("%Y-%m-%d") if hasattr(m, "strftime") else (str(m)[:10] if m else "TBD")
            tdata.append([
                Paragraph(c["nist_id"], cell_style),
                Paragraph((c.get("requirement_text") or "")[:120], cell_style),
                Paragraph((c.get("status") or "").replace("_", " ").title(), cell_style),
                Paragraph(program["org_name"], cell_style),
                Paragraph("TBD", cell_style),
                Paragraph(ms, cell_style),
                Paragraph(c.get("implementation_notes") or "Remediation plan pending.", cell_style),
            ])
        col_w = [0.65 * inch, 2.0 * inch, 1.0 * inch, 1.2 * inch, 0.75 * inch, 0.9 * inch, 2.95 * inch]
        tbl = Table(tdata, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), _colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_colors.white, _colors.HexColor("#f2f4f7")]),
            ("GRID", (0, 0), (-1, -1), 0.3, _colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(tbl)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            rightMargin=0.75 * inch, leftMargin=0.75 * inch,
                            topMargin=1 * inch, bottomMargin=0.75 * inch)
    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    return buf.getvalue()


async def _generate_sprs_bytes(program_id: str, program, controls) -> bytes:
    """Build SPRS xlsx bytes (reuses generate_sprs_xlsx logic, returns raw bytes)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    SPRS_MAX = 110
    sprs = SPRS_MAX
    ssp_ok = True
    for c in controls:
        if not c["is_applicable"]:
            continue
        if c["nist_id"] == "3.12.4" and c["status"] != "fully_implemented":
            ssp_ok = False
        if c["status"] != "fully_implemented" and c["dod_score_value"]:
            sprs -= c["dod_score_value"]
    if not ssp_ok:
        sprs = -203

    DOMAIN_FULL = {
        "AC": "Access Control", "AT": "Awareness & Training", "AU": "Audit & Accountability",
        "CM": "Configuration Mgmt", "IA": "Identification & Auth", "IR": "Incident Response",
        "MA": "Maintenance", "MP": "Media Protection", "PE": "Physical Protection",
        "PS": "Personnel Security", "RA": "Risk Assessment", "CA": "Security Assessment",
        "SC": "System & Comm Protection", "SI": "System & Info Integrity",
    }
    domains: dict = defaultdict(lambda: {"name": "", "total": 0, "implemented": 0, "points_possible": 0, "points_lost": 0})
    for c in controls:
        if not c["is_applicable"]:
            continue
        abbrev = c["family_abbrev"] or "??"
        d = domains[abbrev]
        d["name"] = DOMAIN_FULL.get(abbrev, c["family"] or abbrev)
        d["total"] += 1
        d["points_possible"] += c["dod_score_value"] or 0
        if c["status"] == "fully_implemented":
            d["implemented"] += 1
        else:
            d["points_lost"] += c["dod_score_value"] or 0

    NAVY, BLUE, GREEN, AMBER, RED_C, LTGRAY, WHITE, DARK_RED = (
        "1E3A5F", "2563EB", "16A34A", "D97706", "DC2626", "F1F5F9", "FFFFFF", "991B1B"
    )

    def _fill(h):
        return PatternFill("solid", fgColor=h)

    def _border():
        s = Side(style="thin", color="CBD5E1")
        return Border(left=s, right=s, top=s, bottom=s)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "SPRS Summary"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 36

    ws1.merge_cells("A1:B1")
    c = ws1["A1"]
    _cmmc_level = program.get("cmmc_level", 2)
    c.value = f"DoD CMMC Level {_cmmc_level} — {'SPRS Self-Assessment' if _cmmc_level == 2 else 'Readiness'} Summary"
    c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = _center()
    ws1.row_dimensions[1].height = 30
    ws1.merge_cells("A2:B2")
    c = ws1["A2"]
    c.value = "NIST SP 800-171 Rev 2 | Submit via PIEE Supplier Performance Risk System (SPRS)"
    c.font = Font(italic=True, size=9, color="475569")
    c.alignment = _center()
    ws1.row_dimensions[2].height = 18
    ws1.append([])

    today_str = datetime.utcnow().strftime("%B %d, %Y")

    def _row(label, value, lbold=False, lfill=None, vfill=None):
        r = ws1.max_row + 1
        lc = ws1.cell(r, 1, label)
        vc = ws1.cell(r, 2, value)
        lc.alignment = _left()
        vc.alignment = _left()
        lc.border = _border()
        vc.border = _border()
        if lbold:
            lc.font = Font(bold=True, size=11)
        if lfill:
            lc.fill = _fill(lfill)
        if vfill:
            vc.fill = _fill(vfill)

    _row("Organization", program["org_name"] or "", lbold=True)
    _row("System Name", program.get("system_name") or program["name"] or "")
    _row("CAGE Code", program.get("org_cage_code") or "")
    _row("Assessment Date", today_str)
    _row("Assessor", "Self-Assessment (Internal)")
    ws1.append([])

    score_color = GREEN if sprs >= 70 else (AMBER if sprs >= 0 else DARK_RED)
    r_sprs = ws1.max_row + 1
    lc = ws1.cell(r_sprs, 1, "SPRS SCORE")
    vc = ws1.cell(r_sprs, 2, sprs)
    lc.font = Font(bold=True, size=16, color=WHITE)
    vc.font = Font(bold=True, size=24, color=WHITE)
    lc.fill = _fill(NAVY)
    vc.fill = _fill(score_color)
    lc.alignment = _center()
    vc.alignment = _center()
    ws1.row_dimensions[r_sprs].height = 40
    ws1.append([])

    interp = ("Acceptable — strong compliance posture" if sprs >= 70
              else ("Caution — partial compliance, remediation required" if sprs >= 0
                    else "Critical — SSP not implemented / major gaps (floor: -203)"))
    _row("Score Interpretation", interp, lbold=True,
         vfill=(GREEN if sprs >= 70 else (AMBER if sprs >= 0 else RED_C)))
    _row("Maximum Possible Score", 110)
    _row("Points Deducted", SPRS_MAX - sprs)
    _row("SSP (3.12.4) Implemented", "Yes" if ssp_ok else "No — SPRS floored to -203",
         vfill=("F0FDF4" if ssp_ok else "FEF2F2"))
    ws1.append([])

    piee_hdr = ws1.cell(ws1.max_row + 1, 1, "PIEE Submission Steps")
    piee_hdr.font = Font(bold=True, size=11, color=WHITE)
    piee_hdr.fill = _fill(BLUE)
    piee_hdr.alignment = _center()
    ws1.merge_cells(f"A{ws1.max_row}:B{ws1.max_row}")
    for step, detail in [
        ("Step 1", "Log in to PIEE at https://piee.eb.mil"),
        ("Step 2", "Navigate to: Supplier Performance Risk System (SPRS)"),
        ("Step 3", "Select 'NIST SP 800-171 DoD Assessment' → 'Add Assessment'"),
        ("Step 4", f"Enter your SPRS Score: {sprs}"),
        ("Step 5", "Enter assessment date, system name, and CAGE code"),
        ("Step 6", "Attach your SSP as supporting documentation"),
        ("Step 7", "Submit — visible to DoD contracting officers within 24h"),
        ("Note", "Self-assessment score valid for 3 years per DFARS 252.204-7021"),
    ]:
        r = ws1.max_row + 1
        lc = ws1.cell(r, 1, step)
        vc = ws1.cell(r, 2, detail)
        lc.fill = _fill(LTGRAY)
        lc.font = Font(bold=True, size=10)
        vc.font = Font(size=10)
        lc.alignment = _left()
        vc.alignment = _left()
        lc.border = _border()
        vc.border = _border()

    # Domain breakdown sheet
    ws2 = wb.create_sheet("Domain Breakdown")
    for col, (h, w) in enumerate(zip(
        ["Domain", "Full Name", "Controls", "Implemented", "Pct", "Points Possible", "Points Lost"],
        [8, 30, 10, 13, 9, 15, 13]
    ), 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
        cell = ws2.cell(1, col, h)
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = _fill(NAVY)
        cell.alignment = _center()
        cell.border = _border()
    ws2.row_dimensions[1].height = 22
    for r_idx, (abbrev, d) in enumerate(sorted(domains.items()), 2):
        pct = round(d["implemented"] / d["total"] * 100) if d["total"] else 0
        rf = _fill("F0FDF4") if pct == 100 else (_fill("FEF9C3") if pct >= 50 else _fill("FEF2F2"))
        for col_idx, val in enumerate([abbrev, d["name"], d["total"], d["implemented"], f"{pct}%", d["points_possible"], d["points_lost"]], 1):
            cell = ws2.cell(r_idx, col_idx, val)
            cell.fill = rf
            cell.border = _border()
            cell.font = Font(size=10)
            cell.alignment = _center() if col_idx != 2 else _left()

    # Control detail sheet
    ws3 = wb.create_sheet("Control Detail")
    for col, (h, w) in enumerate(zip(
        ["NIST ID", "CMMC ID", "Domain", "Requirement (truncated)", "Status", "DoD Pts", "Deducted"],
        [10, 14, 8, 52, 22, 9, 10]
    ), 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
        cell = ws3.cell(1, col, h)
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = _fill(NAVY)
        cell.alignment = _center()
        cell.border = _border()
    ws3.row_dimensions[1].height = 22
    ws3.freeze_panes = "A2"
    STATUS_FILLS = {
        "fully_implemented": "F0FDF4", "implementation_begun": "EFF6FF",
        "implementation_planned": "FEFCE8", "not_yet_addressed": "FEF2F2",
        "not_yet_assessed": "F8FAFC", "not_applicable": "F1F5F9",
    }
    for r_idx, c in enumerate(controls, 2):
        if not c["is_applicable"]:
            continue
        deducted = c["dod_score_value"] if c["status"] != "fully_implemented" else 0
        req_text = c.get("requirement_text") or ""
        req = req_text[:120] + "…" if len(req_text) > 120 else req_text
        rf = _fill(STATUS_FILLS.get(c["status"] or "", "FFFFFF"))
        for col_idx, val in enumerate([
            c["nist_id"], c.get("cmmc_id") or "", c.get("family_abbrev") or "",
            req, (c.get("status") or "").replace("_", " ").title(),
            c.get("dod_score_value") or 0, deducted
        ], 1):
            cell = ws3.cell(r_idx, col_idx, val)
            cell.fill = rf
            cell.border = _border()
            cell.font = Font(size=9)
            cell.alignment = _center() if col_idx not in (4, 5) else _left()
            if col_idx == 7 and deducted:
                cell.font = Font(size=9, color=DARK_RED, bold=True)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
