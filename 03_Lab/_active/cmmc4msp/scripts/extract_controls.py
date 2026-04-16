"""
extract_controls.py

Reads 'CMMC Dashboard.xlsx' and generates
'postgres/migrations/002_controls_seed.sql' with INSERT statements
for all parent controls and sub-objectives found in the SSP sheet.

Sheet sources:
  - "SSP - DoD SPRS & FAR and Above"  (row 2 = headers, row 3+ = data)
  - "Potential Assessment Consid."     (row 1 = headers, row 2+ = data)

Output table: cmmc_main.control_definitions
"""

import uuid
import re
import os
import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)

EXCEL_PATH = os.path.join(PROJECT_DIR, "files", "CMMC Dashboard.xlsx")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "postgres", "migrations")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "002_controls_seed.sql")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
UUID_NAMESPACE = uuid.NAMESPACE_DNS


def make_uuid(nist_id: str) -> str:
    """Deterministic UUID from NIST ID string."""
    return str(uuid.uuid5(UUID_NAMESPACE, nist_id.strip()))


def esc(value) -> str:
    """
    Escape a value for SQL single-quoted string literal.
    Returns NULL (unquoted) for None/empty.
    """
    if value is None:
        return "NULL"
    s = str(value).strip()
    if not s:
        return "NULL"
    # Escape single quotes
    s = s.replace("'", "''")
    return f"'{s}'"


def esc_bool(value: bool) -> str:
    return "TRUE" if value else "FALSE"


def esc_int(value) -> str:
    if value is None:
        return "NULL"
    try:
        return str(int(value))
    except (ValueError, TypeError):
        return "NULL"


def family_abbrev_from_cmmc(cmmc_id) -> str:
    """Extract family abbreviation from CMMC control ID.
    e.g. 'AC.L1-3.1.1' -> 'AC'
    """
    if not cmmc_id:
        return ""
    s = str(cmmc_id).strip()
    m = re.match(r'^([A-Z]+)\.', s)
    if m:
        return m.group(1)
    return ""


def parent_nist_id(nist_id: str) -> str:
    """Strip sub-objective suffix to get parent NIST ID.
    e.g. '3.13.1[a]' -> '3.13.1'
    """
    return re.sub(r'\[.*\]$', '', nist_id).strip()


def diy_type_map(raw) -> str:
    if not raw:
        return "hybrid"
    s = str(raw).strip().lower()
    if s == "outsource":
        return "outsource"
    if s == "diy":
        return "diy"
    return "hybrid"


# ---------------------------------------------------------------------------
# Load assessment proof guidance from "Potential Assessment Consid." sheet
# ---------------------------------------------------------------------------
def load_proof_guidance(wb) -> dict:
    """
    Returns dict keyed by NIST ID -> acceptable_proof_guidance string.
    Combines Examine (col F) and Interview (col G) columns.
    """
    ws = wb["Potential Assessment Consid."]
    guidance = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nist_id = row[3]  # col D
        examine = row[5]  # col F
        interview = row[6]  # col G
        if not nist_id:
            continue
        nist_id = str(nist_id).strip()
        parts = []
        if examine and str(examine).strip():
            parts.append(f"EXAMINE: {str(examine).strip()}")
        if interview and str(interview).strip():
            parts.append(f"INTERVIEW: {str(interview).strip()}")
        if parts:
            guidance[nist_id] = "\n\n".join(parts)
    return guidance


# ---------------------------------------------------------------------------
# Parse SSP sheet
# ---------------------------------------------------------------------------
def parse_ssp(wb, guidance: dict) -> tuple[list, list]:
    """
    Returns (parent_rows, sub_rows) — each is a list of dicts ready for SQL.

    SSP columns (0-indexed from row 3+):
      0  A  far_above_phase
      1  B  far_above_sort_order
      2  C  diy_type raw
      3  D  family
      4  E  nist_id
      5  F  nist_sort_order
      6  G  cmmc_id
      7  H  dod_score_value
      8  I  requirement_text
      9  J  assessment_objective
     10  K  dod_comment
     11  L  (client impl text — skip)
     12  M  implementation_status (skip for seed)
     13  N  target date (skip)
     14  O  score_impact (skip)
    """
    ws = wb["SSP - DoD SPRS & FAR and Above"]

    parents = []
    subs = []

    for row in ws.iter_rows(min_row=3, max_col=15, values_only=True):
        nist_id_raw = row[4]
        if nist_id_raw is None:
            continue
        nist_id = str(nist_id_raw).strip()
        if not nist_id:
            continue

        far_phase = str(row[0]).strip() if row[0] is not None else ""
        far_sort = str(row[1]).strip() if row[1] is not None else ""
        diy_raw = row[2]
        family = str(row[3]).strip() if row[3] is not None else ""
        nist_sort = str(row[5]).strip() if row[5] is not None else ""
        cmmc_id = str(row[6]).strip() if row[6] is not None else ""
        dod_score = row[7]
        requirement = row[8]
        assessment_obj = row[9]
        dod_comment = row[10]

        abbrev = family_abbrev_from_cmmc(cmmc_id)
        is_basic = far_phase == "1"
        diy = diy_type_map(diy_raw)

        if "[" in nist_id:
            # Sub-objective
            parent_id_str = parent_nist_id(nist_id)
            rec = {
                "id": make_uuid(nist_id),
                "nist_id": nist_id,
                "nist_sort_order": nist_sort,
                "cmmc_id": cmmc_id,
                "family": family,
                "family_abbrev": abbrev,
                "far_above_phase": far_phase,
                "far_above_sort_order": far_sort,
                "diy_type": diy,
                "is_objective": True,
                "parent_control_id": make_uuid(parent_id_str),
                "dod_score_value": None,
                "requirement_text": None,
                "assessment_objective": assessment_obj,
                "dod_comment": dod_comment,
                "acceptable_proof_guidance": guidance.get(parent_id_str),
                "is_basic": is_basic,
            }
            subs.append(rec)
        else:
            # Parent control — include even if dod_score == 0
            rec = {
                "id": make_uuid(nist_id),
                "nist_id": nist_id,
                "nist_sort_order": nist_sort,
                "cmmc_id": cmmc_id,
                "family": family,
                "family_abbrev": abbrev,
                "far_above_phase": far_phase,
                "far_above_sort_order": far_sort,
                "diy_type": diy,
                "is_objective": False,
                "parent_control_id": None,
                "dod_score_value": dod_score,
                "requirement_text": requirement,
                "assessment_objective": assessment_obj,
                "dod_comment": dod_comment,
                "acceptable_proof_guidance": guidance.get(nist_id),
                "is_basic": is_basic,
            }
            parents.append(rec)

    return parents, subs


# ---------------------------------------------------------------------------
# Generate SQL
# ---------------------------------------------------------------------------
COLUMNS = [
    "id",
    "nist_id",
    "nist_sort_order",
    "cmmc_id",
    "family",
    "family_abbrev",
    "far_above_phase",
    "far_above_sort_order",
    "diy_type",
    "is_objective",
    "parent_control_id",
    "dod_score_value",
    "requirement_text",
    "assessment_objective",
    "dod_comment",
    "acceptable_proof_guidance",
    "is_basic",
]


def row_to_values(r: dict) -> str:
    """Convert a record dict to a SQL VALUES tuple string."""
    parts = [
        f"'{r['id']}'::uuid",                        # id
        esc(r["nist_id"]),                            # nist_id
        esc(r["nist_sort_order"]),                    # nist_sort_order
        esc(r["cmmc_id"]),                            # cmmc_id
        esc(r["family"]),                             # family
        esc(r["family_abbrev"]),                      # family_abbrev
        esc(r["far_above_phase"]),                    # far_above_phase
        esc(r["far_above_sort_order"]),               # far_above_sort_order
        esc(r["diy_type"]),                           # diy_type
        esc_bool(r["is_objective"]),                  # is_objective
        (f"'{r['parent_control_id']}'::uuid"          # parent_control_id
         if r["parent_control_id"] else "NULL"),
        esc_int(r["dod_score_value"]),                # dod_score_value
        esc(r["requirement_text"]),                   # requirement_text
        esc(r["assessment_objective"]),               # assessment_objective
        esc(r["dod_comment"]),                        # dod_comment
        esc(r["acceptable_proof_guidance"]),          # acceptable_proof_guidance
        esc_bool(r["is_basic"]),                      # is_basic
    ]
    return "    (" + ",\n     ".join(parts) + ")"


def generate_sql(parents: list, subs: list) -> str:
    col_list = ",\n    ".join(COLUMNS)
    lines = []

    lines.append("-- =============================================================")
    lines.append("-- 002_controls_seed.sql")
    lines.append("-- Auto-generated by scripts/extract_controls.py")
    lines.append(f"-- Parent controls : {len(parents)}")
    lines.append(f"-- Sub-objectives  : {len(subs)}")
    lines.append(f"-- Total records   : {len(parents) + len(subs)}")
    lines.append("-- =============================================================")
    lines.append("")
    lines.append("SET search_path TO cmmc_main, public;")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # ------------------------------------------------------------------
    # Parent controls first (sub-objectives reference them)
    # ------------------------------------------------------------------
    lines.append("-- -------------------------------------------------------")
    lines.append(f"-- Parent controls ({len(parents)} rows)")
    lines.append("-- -------------------------------------------------------")
    lines.append(f"INSERT INTO control_definitions (")
    lines.append(f"    {col_list}")
    lines.append(") VALUES")

    parent_values = [row_to_values(r) for r in parents]
    lines.append(",\n".join(parent_values))
    lines.append("ON CONFLICT (id) DO NOTHING;")
    lines.append("")

    # ------------------------------------------------------------------
    # Sub-objectives
    # ------------------------------------------------------------------
    lines.append("-- -------------------------------------------------------")
    lines.append(f"-- Sub-objectives ({len(subs)} rows)")
    lines.append("-- -------------------------------------------------------")
    lines.append(f"INSERT INTO control_definitions (")
    lines.append(f"    {col_list}")
    lines.append(") VALUES")

    sub_values = [row_to_values(r) for r in subs]
    lines.append(",\n".join(sub_values))
    lines.append("ON CONFLICT (id) DO NOTHING;")
    lines.append("")

    lines.append("COMMIT;")
    lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    print("Loading proof guidance...")
    guidance = load_proof_guidance(wb)
    print(f"  Guidance entries loaded: {len(guidance)}")

    print("Parsing SSP sheet...")
    parents, subs = parse_ssp(wb, guidance)
    print(f"  Parent controls  : {len(parents)}")
    print(f"  Sub-objectives   : {len(subs)}")
    print(f"  Total records    : {len(parents) + len(subs)}")

    print(f"Generating SQL -> {OUTPUT_PATH}")
    sql = generate_sql(parents, subs)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(sql)

    size_kb = os.path.getsize(OUTPUT_PATH) / 1024
    print(f"  Written: {size_kb:.1f} KB")
    print("Done.")


if __name__ == "__main__":
    main()
